"""
Excel export functionality
"""

import logging
from typing import List, Dict
from datetime import datetime
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export job data to Excel file"""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_jobs_to_excel(self, jobs: List[Dict], filename: str = None) -> str:
        """Export jobs to Excel file"""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl is not installed. Cannot export to Excel.")
            return None
        
        if not jobs:
            logger.warning("No jobs to export")
            return None
        
        if filename is None:
            filename = f"job_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "German Job Postings"
            
            # Define headers
            headers = [
                "Company Name",
                "Job Title",
                "Job URL",
                "Posted Date",
                "Job Board",
                "Location",
                "Job Level",
                "Keywords Match",
                "Rank"
            ]
            
            # Add headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Add data rows
            data_font = Font(size=11)
            data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            for row_num, job in enumerate(jobs, 2):
                row_data = [
                    job.get('company_name', ''),
                    job.get('job_title', ''),
                    job.get('job_url', ''),
                    job.get('posted_date', ''),
                    job.get('job_board', ''),
                    job.get('location', ''),
                    job.get('job_level', ''),
                    job.get('keyword_matches', ''),
                    job.get('rank', '')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # Make URLs clickable
                    if col_num == 3 and value:  # Job URL column
                        cell.hyperlink = value
                        cell.font = Font(color="0563C1", underline="single")
            
            # Adjust column widths
            column_widths = {
                'A': 20,  # Company Name
                'B': 35,  # Job Title
                'C': 40,  # Job URL
                'D': 15,  # Posted Date
                'E': 15,  # Job Board
                'F': 20,  # Location
                'G': 15,  # Job Level
                'H': 30,  # Keywords Match
                'I': 8,   # Rank
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws['A1'] = "Job Search Summary"
            summary_ws['A1'].font = Font(bold=True, size=14)
            
            summary_ws['A3'] = "Total Jobs Found:"
            summary_ws['B3'] = len(jobs)
            
            summary_ws['A4'] = "Search Date:"
            summary_ws['B4'] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            
            # Count by board
            board_counts = {}
            for job in jobs:
                board = job.get('job_board', 'Unknown')
                board_counts[board] = board_counts.get(board, 0) + 1
            
            summary_ws['A6'] = "Jobs by Board:"
            row = 7
            for board, count in sorted(board_counts.items(), key=lambda x: x[1], reverse=True):
                summary_ws[f'A{row}'] = board
                summary_ws[f'B{row}'] = count
                row += 1
            
            # Adjust summary sheet columns
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 20
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"Exported {len(jobs)} jobs to {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            return None


class CSVExporter:
    """Export job data to CSV file (fallback if openpyxl not available)"""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_jobs_to_csv(self, jobs: List[Dict], filename: str = None) -> str:
        """Export jobs to CSV file"""
        import csv
        
        if not jobs:
            logger.warning("No jobs to export")
            return None
        
        if filename is None:
            filename = f"job_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            fieldnames = [
                "company_name",
                "job_title",
                "job_url",
                "posted_date",
                "job_board",
                "location",
                "job_level",
                "keyword_matches",
                "rank"
            ]
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for job in jobs:
                    writer.writerow({
                        'company_name': job.get('company_name', ''),
                        'job_title': job.get('job_title', ''),
                        'job_url': job.get('job_url', ''),
                        'posted_date': job.get('posted_date', ''),
                        'job_board': job.get('job_board', ''),
                        'location': job.get('location', ''),
                        'job_level': job.get('job_level', ''),
                        'keyword_matches': job.get('keyword_matches', ''),
                        'rank': job.get('rank', '')
                    })
            
            logger.info(f"Exported {len(jobs)} jobs to {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            return None
